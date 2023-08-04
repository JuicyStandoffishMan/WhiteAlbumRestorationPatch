import os
import re
import openpyxl

# Directory paths
eng_dir = 'eng'
jap_dir = 'jap'
out_dir = 'merged'

if not os.path.exists(out_dir):
    os.makedirs(out_dir)

# Iterate over English Excel files
for filename in os.listdir(eng_dir):
    if filename.endswith('.xlsx'):
        eng_path = os.path.join(eng_dir, filename)
        jap_path = os.path.join(jap_dir, filename)

        # Open both English and Japanese Excel files
        eng_wb = openpyxl.load_workbook(eng_path)
        jap_wb = openpyxl.load_workbook(jap_path)

        eng_ws = eng_wb.active
        jap_ws = jap_wb.active

        # Check row count
        if eng_ws.max_row != jap_ws.max_row:
            print(f"Warning: Mismatch in number of rows for {filename}")
            # continue

        # Process rows
        japanese_lines = [jap_ws.cell(row=row, column=5).value for row in range(1, jap_ws.max_row + 1)]
        for row in range(1, eng_ws.max_row + 1):
            jap_cell = jap_ws.cell(row=row, column=5)  # Column E
            eng_cell = eng_ws.cell(row=row, column=5)  # Column E
            eng_cell.value = jap_cell.value
            

        # Write modified English Excel file to 'out' directory
        out_path = os.path.join(out_dir, filename)
        eng_wb.save(out_path)

        print(f"Processed {filename}")