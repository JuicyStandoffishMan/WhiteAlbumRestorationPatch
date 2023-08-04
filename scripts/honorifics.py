import os
import re
import openpyxl

# Mapping of English first names to Japanese characters and their corresponding last names.
name_mapping = {
    "Touya Fujii": {
        "first_name": "冬弥",
        "last_name": "藤井"
    },
    "Haruka Kawashima": {
        "first_name": "はるか",
        "last_name": "河島"
    },
    "Sayoko Kisaragi": {
        "first_name": "小夜子",
        "last_name": "如月"
    },
    "Mana Mizuki": {
        "first_name": "マナ",
        "last_name": "観月"
    },
    "Yuki Morikawa": {
        "first_name": "由綺",
        "last_name": "森川"
    },
    "Rina Ogata": {
        "first_name": "理奈",
        "last_name": "緒方"
    },
    "Misaki Sawakura": {
        "first_name": "美咲",
        "last_name": "澤倉"
    },
    "Yayoi Shinozuka": {
        "first_name": "弥生",
        "last_name": "篠塚"
    },
    "Akira Nanase": {
        "first_name": "彰",
        "last_name": "七瀬"
    },
    "Eiji Ogata": {
        "first_name": "英二",
        "last_name": "緒方"
    }
}

# Mapping of Japanese honorifics.
honorific_mapping = {
    "さん": "-san",
    "くん": "-kun",
    "君": "-kun",
    "ちゃん": "-chan",
    "さま": "-sama",
    "せんせい": "-sensei",
    "先生": "-sensei",
    "ねえさん" : "-neesan",
}

def add_honorifics(english_line, japanese_lines, current_line):
    for eng_full_name, name_info in name_mapping.items():
        eng_occurrences_full = english_line.count(eng_full_name)
        eng_occurrences_first = english_line.count(eng_full_name.split(' ')[0])
        eng_occurrences_last = english_line.count(eng_full_name.split(' ')[1])
        
        if eng_occurrences_full > 0 or eng_occurrences_first > 0 or eng_occurrences_last > 0:
            lines_to_check = [current_line]  # Start with the current Japanese line
            if current_line > 0:  # Add previous line if available
               lines_to_check.append(current_line - 1)
            if current_line < len(japanese_lines) - 1:  # Add next line if available
                lines_to_check.append(current_line + 1)

            for line_number in lines_to_check:
                japanese_line = japanese_lines[line_number]
                if japanese_line is None:
                    continue
                matched_honorific = None

                jap_occurrences_first = 0
                jap_occurrences_last = 0
                for honorific_jap, honorific_eng in honorific_mapping.items():
                    # Check for full name first
                    full_name_patterns = [name_info['first_name'] + name_info['last_name'] + honorific_jap,
                                          name_info['last_name'] + name_info['first_name'] + honorific_jap]
                    jap_occurrences_full = sum(japanese_line.count(pattern) for pattern in full_name_patterns)

                    # Check for first and last names separately
                    partial_name_patterns = [name_info['first_name'] + honorific_jap, name_info['last_name'] + honorific_jap]
                    jap_occurrences_first = japanese_line.count(name_info['first_name'] + honorific_jap)
                    jap_occurrences_last = japanese_line.count(name_info['last_name'] + honorific_jap)

                    if (jap_occurrences_full == eng_occurrences_full and jap_occurrences_full > 0) or \
                       ((jap_occurrences_first == eng_occurrences_first or jap_occurrences_last == eng_occurrences_first) and \
                       (jap_occurrences_last == eng_occurrences_last or jap_occurrences_first == eng_occurrences_last) and (jap_occurrences_first > 0 or jap_occurrences_last > 0)):
                        if matched_honorific is None:
                            matched_honorific = honorific_eng
                            break
                        elif matched_honorific != honorific_eng:
                            matched_honorific = None
                            break

                if matched_honorific is not None:
                    if eng_occurrences_full > 0:
                        english_line = english_line.replace(eng_full_name, eng_full_name + matched_honorific)
                    elif eng_occurrences_first > 0:
                        if jap_occurrences_first > 0:
                            english_line = english_line.replace(eng_full_name.split(' ')[0], eng_full_name.split(' ')[0] + matched_honorific)
                        else:
                            english_line = english_line.replace(eng_full_name.split(' ')[0], eng_full_name.split(' ')[1] + matched_honorific)
                    elif eng_occurrences_last > 0:
                        if jap_occurrences_last > 0:
                            english_line = english_line.replace(eng_full_name.split(' ')[1], eng_full_name.split(' ')[1] + matched_honorific)
                        else:
                            english_line = english_line.replace(eng_full_name.split(' ')[1], eng_full_name.split(' ')[0] + matched_honorific)
                    break
                continue

                # Finding the Japanese name and character after the name
                pattern = name_info['first_name'] + name_info['last_name'] + r'.'  # Assuming full name; adjust as needed
                match = re.search(pattern, japanese_line)
                if match:
                    context = match.group(0)  # Include the character after the name
                    print(f"Ambiguity found for {eng_full_name}. Unable to determine the honorific. Context: {context}")
                else:
                    print(f"Ambiguity found for {eng_full_name}. Unable to determine the honorific.")

    return english_line



def add_missing_spaces(text):
    # Find occurrences where a character is followed by an alphanumeric character without a space between
    corrected_text = re.sub(r'([.!?])([A-Za-z])', r'\1 \2', text)
    return corrected_text

def remove_miss(english_line, current_line):
    if "Miss " in english_line:
        english_line = english_line.replace("Miss ", "")
        print(f"Removed 'Miss ' from line {current_line}: " + english_line)
    if "Mr. " in english_line:
        english_line = english_line.replace("Mr. ", "")
        print(f"Removed 'Mr. ' from line {current_line}: " + english_line)
    if "Mrs. " in english_line:
        english_line = english_line.replace("Mrs. ", "")
        print(f"Removed 'Mrs. ' from line {current_line}: " + english_line)
    if "Ms. " in english_line:
        english_line = english_line.replace("Ms. ", "")
        print(f"Removed 'Ms. ' from line {current_line}: " + english_line)
    if "Prof. " in english_line:
        english_line = english_line.replace("Prof. ", "")
        print(f"Removed 'Prof. ' from line {current_line}: " + english_line)
    if "Professor. " in english_line:
        english_line = english_line.replace("Prof. ", "")
        print(f"Removed 'Professor. ' from line {current_line}: " + english_line)

    return english_line

# Directory paths
in_dir = 'merged'
out_dir = 'merged'

# Iterate over English Excel files
for filename in os.listdir(in_dir):
    if filename.endswith('.xlsx'):
        eng_path = os.path.join(in_dir, filename)

        # Open both English and Japanese Excel files
        eng_wb = openpyxl.load_workbook(eng_path)

        eng_ws = eng_wb.active

        # Process rows
        japanese_lines = [eng_ws.cell(row=row, column=5).value for row in range(1, eng_ws.max_row + 1)]
        for row in range(1, eng_ws.max_row + 1):
            eng_cell = eng_ws.cell(row=row, column=7)  # Column G
            edited_cell = eng_ws.cell(row=row, column=6)  # Column F
            edited_cell.value = add_honorifics(remove_miss(add_missing_spaces(eng_cell.value), row - 1), japanese_lines, row - 1)
            if edited_cell.value == eng_cell.value:
                edited_cell.value = None

        # Write modified English Excel file to 'out' directory
        out_path = eng_path
        eng_wb.save(out_path)

        print(f"Processed {filename}")