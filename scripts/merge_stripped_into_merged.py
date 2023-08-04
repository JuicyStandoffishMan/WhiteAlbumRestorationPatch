import os
import re
import openpyxl

# Directory paths
merged_dir = 'merged'
stripped_dir = 'stripped'

for filename in os.listdir(merged_dir):
    if filename.endswith('.xlsx'):
        merged_path = os.path.join(merged_dir, filename)
        stripped_path = os.path.join(stripped_dir, filename)

        # Open both English and Japanese Excel files
        eng_wb = openpyxl.load_workbook(merged_path)
        eng_ws = eng_wb.active

        stripped_wb = openpyxl.load_workbook(stripped_path)
        stripped_ws = stripped_wb.active

        # Process rows
        for row in range(1, eng_ws.max_row + 1):
            eng_ws.cell(row=row, column=6).value = stripped_ws.cell(row=row, column=6).value  # Column F

        # Write modified English Excel file to 'out' directory
        eng_wb.save(merged_path)

        print(f"Processed {filename}")