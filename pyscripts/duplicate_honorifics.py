import os
import re
import openpyxl

# Mapping of English first names to Japanese characters and their corresponding last names.
name_mapping = {
    "Touya Fujii": {
        "first_name": "冬弥",
        "last_name": "藤井"
    },
    "Haruka Kawashima": {
        "first_name": "はるか",
        "last_name": "河島"
    },
    "Sayoko Kisaragi": {
        "first_name": "小夜子",
        "last_name": "如月"
    },
    "Mana Mizuki": {
        "first_name": "マナ",
        "last_name": "観月"
    },
    "Yuki Morikawa": {
        "first_name": "由綺",
        "last_name": "森川"
    },
    "Rina Ogata": {
        "first_name": "理奈",
        "last_name": "緒方"
    },
    "Misaki Sawakura": {
        "first_name": "美咲",
        "last_name": "澤倉"
    },
    "Yayoi Shinozuka": {
        "first_name": "弥生",
        "last_name": "篠塚"
    },
    "Akira Nanase": {
        "first_name": "彰",
        "last_name": "七瀬"
    },
    "Eiji Ogata": {
        "first_name": "英二",
        "last_name": "緒方"
    }
}

# Mapping of Japanese honorifics.
honorific_mapping = {
    "さん": "-san",
    "くん": "-kun",
    "君": "-kun",
    "ちゃん": "-chan",
    "さま": "-sama",
    "せんせい": "-sensei",
    "先生": "-sensei",
    "ねえさん" : "-neesan",
}

def fix_honorifics(english_line, japanese_lines, current_line):
    if english_line is None:
        return english_line
    # Remove duplicate honorifics from the english line, like -san-san
    english_line = re.sub(r'(-san)+', r'-san', english_line)
    english_line = re.sub(r'(-kun)+', r'-kun', english_line)
    english_line = re.sub(r'(-chan)+', r'-chan', english_line)
    english_line = re.sub(r'(-sama)+', r'-sama', english_line)
    english_line = re.sub(r'(-sensei)+', r'-sensei', english_line)
    english_line = re.sub(r'(-neesan)+', r'-neesan', english_line)

    return english_line



# Directory paths
in_dir = 'merged'
out_dir = 'merged'

# Iterate over English Excel files
for filename in os.listdir(in_dir):
    if filename.endswith('.xlsx'):
        eng_path = os.path.join(in_dir, filename)

        # Open both English and Japanese Excel files
        eng_wb = openpyxl.load_workbook(eng_path)

        eng_ws = eng_wb.active

        # Process rows
        japanese_lines = [eng_ws.cell(row=row, column=5).value for row in range(1, eng_ws.max_row + 1)]
        for row in range(1, eng_ws.max_row + 1):
            edited_cell = eng_ws.cell(row=row, column=6)  # Column F
            original = edited_cell.value
            edited_cell.value = fix_honorifics(edited_cell.value, japanese_lines, row - 1)
            if edited_cell.value != original:
                print(f"Processed {filename}: {original} -> {edited_cell.value}")
                out_path = eng_path
                eng_wb.save(out_path)
            