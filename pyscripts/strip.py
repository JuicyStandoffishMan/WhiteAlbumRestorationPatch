import os
import re
import openpyxl

# Directory paths
in_dir = 'merged'
out_dir = 'stripped'

if not os.path.exists(out_dir):
    os.makedirs(out_dir)
for filename in os.listdir(in_dir):
    if filename.endswith('.xlsx'):
        eng_path = os.path.join(in_dir, filename)

        # Open both English and Japanese Excel files
        eng_wb = openpyxl.load_workbook(eng_path)

        eng_ws = eng_wb.active

        # Process rows
        japanese_lines = [eng_ws.cell(row=row, column=5).value for row in range(1, eng_ws.max_row + 1)]
        for row in range(1, eng_ws.max_row + 1):
            eng_cell = eng_ws.cell(row=row, column=7)  # Column G
            jp_cell = eng_ws.cell(row=row, column=5)  # Column F
            eng_cell.value = None
            jp_cell.value = None

        # Write modified English Excel file to 'out' directory
        out_path = os.path.join(out_dir, filename)
        eng_wb.properties.creator = 'Anonymous'
        eng_wb.properties.lastModifiedBy = 'Anonymous'

        # Remove defined names (including named ranges)
        for name in eng_wb.defined_names:
            del eng_wb.defined_names[name.name]

        eng_wb.custom_properties = {}

        eng_wb.save(out_path)

        print(f"Processed {filename}")